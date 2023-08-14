import logging
from os import path
from openpyxl import load_workbook
from selenium import webdriver
from scraper.account import Account
from scraper.constants import Constants as constants
from scraper.person import Person
from scraper.search import Search

# Group profiles scraper
crome_path = path.abspath("chrome_driver/chromedriver")
driver = webdriver.Chrome(crome_path)
account = Account(driver)
account.login(constants.email, constants.password)
# Get search parameters
wb = load_workbook('search_results.xlsx')
ws_search = wb["Search_parameters"]
page_limit = ws_search['D2'].value
results_limit = ws_search['E2'].value
my_search = Search(
    driver, ws_search['A2'].value,  ws_search['B2'].value, ws_search['C2'].value, pages_limit=page_limit)
# Start search
my_search.search()
urls = my_search.results_url_array
# Get persons data
result_persons = []
for url in urls:
    try:
        my_person = Person(driver, person_url=url)
        result_persons.append(my_person)
        # page limit
        if len(result_persons) >= results_limit:
            break
    except:
        logging.ERROR("Person data error url: " + url)
# Save data to excel
# Create a new sheet
ws = wb.create_sheet(title="LinkedIn data")
# excel titles
ws.append(["name", "job title", "location", "connections", "about",
          "experiences", "educations", "accomplishments", "profile url"])
for person in result_persons:
    data = []
    data.append(person.name)
    data.append(person.job_title)
    data.append(person.location)
    data.append(person.connections)
    data.append("{about}".format(about=person.about))
    data.append("{experiences}".format(experiences=person.experiences))
    data.append("{educations}".format(educations=person.educations))
    data.append("{accomplishment}".format(
        accomplishment=person.accomplishments))
    data.append("{url}".format(url=person.person_url))
    ws.append(data)
# Save data
wb.save('search_results.xlsx')
